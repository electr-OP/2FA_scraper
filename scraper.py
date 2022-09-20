from concurrent.futures import thread
import os
import ast
import random
import time
import zipfile
import pandas as pd
from fake_useragent import UserAgent
from pytz import timezone
from openpyxl import load_workbook
from multiprocessing.pool import ThreadPool, Pool
import concurrent.futures
#selenium libraries
from selenium.webdriver.common import keys
import undetected_chromedriver as uc
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException   
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from read_email import email_code
from testtt2 import get_col_vlaues
from pdf2excel import *
from helper import *
# from new_read_email import *


lisst = pd.read_excel('new_Opeyemi email list with generated email.xlsx')
u = lisst['LOGIN original Gmail'].tolist()
ps = lisst['PASSWORD'].tolist()

def delay ():
    time.sleep(random.randint(2,3))

with open(r'statics\text\premium_http_proxies.txt', 'r') as p:
    lister = p.readlines()

def current_alias_func():
    with open(r'statics\text\email_list1.txt', 'r') as p:
        all_cred = p.readlines()
    # cuurent_alias= 'ALIAS'+str(random.randint(0,999999))
    r = random.choice(u)
    ind = u.index(r)
    curr_u = r
    curr_p = ps[ind]

    if curr_u in all_cred:
        return current_alias_func()
    with open(r'statics\text\email_list1.txt', 'a') as p:
        p.write(f"{curr_u}\n")
    return curr_u, curr_p

# def current_alias_func():
#     import string    
#     import random # define the random module  
#     S = 10  # number of characters in the string.  
#     # call random.choices() string module to find the string in Uppercase + numeric data.  
#     ran = ''.join(random.choices(string.ascii_uppercase + string.digits, k = S))    
#     return str(ran)

# def select_domain():
#     import random
#     domains = ['wkjfher.xyz', 'pqiqef.xyz', 'eirufw.xyz']
#     return random.choice(domains)


# account credentials


# username1 = "tim2@uernwq.xyz"
# username2 = "lukereed@uernwq.xyz"
# username3 = "barnes@uernwq.xyz"
# password = "quebec01"


def get_chromedriver(lister_port, thread, use_proxy=False, user_agent=None):
    try:
        PROXY_HOST = 'ca.smartproxy.com'  # rotating proxy or host
        PROXY_PORT = lister_port # port '20000' 
        PROXY_USER = 'user-2fascraper-sessionduration-10' # username 'user-2fascraper' 
        PROXY_PASS = 'test123' # password


        manifest_json = """
        {
            "version": "1.0.0",
            "manifest_version": 2,
            "name": "Chrome Proxy",
            "permissions": [
                "proxy",
                "tabs",
                "unlimitedStorage",
                "storage",
                "<all_urls>",
                "webRequest",
                "webRequestBlocking"
            ],
            "background": {
                "scripts": ["background.js"]
            },
            "minimum_chrome_version":"22.0.0"
        }
        """

        background_js = """
        var config = {
                mode: "fixed_servers",
                rules: {
                singleProxy: {
                    scheme: "http",
                    host: "%s",
                    port: parseInt(%s)
                },
                bypassList: ["localhost"]
                }
            };

        chrome.proxy.settings.set({value: config, scope: "regular"}, function() {});

        function callbackFn(details) {
            return {
                authCredentials: {
                    username: "%s",
                    password: "%s"
                }
            };
        }

        chrome.webRequest.onAuthRequired.addListener(
                    callbackFn,
                    {urls: ["<all_urls>"]},
                    ['blocking']
        );
        """ % (PROXY_HOST, PROXY_PORT, PROXY_USER, PROXY_PASS)



        # path = os.path.dirname(os.path.abspath(__file__))
        chrome_options = uc.ChromeOptions()
        chrome_options.add_argument('--no-first-run --no-service-autorun --password-store=basic')
        chrome_options.add_argument(f'--user-data-dir={os.getcwd()}\\profiles\\profile{str(thread)}')
        chrome_options.add_experimental_option('prefs', {
        "download.default_directory": f'{os.getcwd()}\\statics\\pdfs{thread}', #Change default directory for downloads
        "download.prompt_for_download": False, #To auto download the file
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome
        })
        if use_proxy:
            pluginfile = 'proxy_auth_plugin.zip'

            with zipfile.ZipFile(pluginfile, 'w') as zp:
                zp.writestr("manifest.json", manifest_json)
                zp.writestr("background.js", background_js)
            chrome_options.add_extension(pluginfile)
        if user_agent:
            chrome_options.add_argument('--user-agent=%s' % user_agent)
        
        driver = uc.Chrome(chrome_options=chrome_options)
        driver.maximize_window()
        return driver

    except Exception as chromedriver_error:
        print(chromedriver_error)

def main(inputs, output, thread):
    
    links = pd.read_excel('statics\excel\SJSR_Granby.xlsx')
    nums = inputs
    count = 1
    # out = pd.read_excel(output)
    out = get_sheetcol_values(output, 1)[1:]
    for liste in lister:
        liste = liste.split(':')[1]
        for i in range(10):
            combined = []
            new = False
            while new == False:
                # if nums[count] not in out['Lot'].tolist():
                if nums[count] not in out:
                    new = True
                else:
                    count+=1
            
            try:
                link1 = links['Portal'][0]
                link11 = 'https://e-services.acceo.com/immosoft/controller/ImmoNetPub/U4051/trouverParCadastre?session_id=64BD8A349A623D00C581BF094361F20C&fourn_seq=47&init_mapping=&platform_name=Windows&browser_name=Chrome&browser_version=97&IdImmeuble=&NoMatricule=&FindMode=1&NomRue=&cd_secteur=&NomComposeRue=&NoCivique=&RechNomRue=&Local=&ty_rapport=TAXATION&g-recaptcha-response='
                link2 = 'https://e-services.acceo.com/immosoft/controller/ImmoNetPub/U4051/trouverParAdresse?init_mapping=&language=en&fourn_seq=63'
                link22 = 'https://e-services.acceo.com/immosoft/controller/ImmoNetPub/U4051/trouverParCadastre?session_id=64BD8A349A623D00C581BF094361F20C&fourn_seq=63&init_mapping=&platform_name=Windows&browser_name=Chrome&browser_version=97&IdImmeuble=&NoMatricule=&FindMode=1&NomRue=&cd_secteur=&NomComposeRue=&NoCivique=&RechNomRue=&Local=&ty_rapport=TAXATION&g-recaptcha-response='
                link3 = 'https://e-services.acceo.com/immosoft/controller/ImmoNetPub/U4051/trouverParAdresse?init_mapping=&language=fr&fourn_seq=224'
                driver = get_chromedriver(liste, thread, use_proxy=True)
                # driver.set_page_load_timeout(30)
                delay()
                # driver.get('http://whatismyip.com')
                
                driver.get(link1)
                for r in range(3):
                    driver.get(link1)
                    try:
                        WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, "//li[@id='Find-3']/a")))
                        break
                    except:
                        continue
                delay()
                try:
                    WebDriverWait(driver,2).until(EC.presence_of_element_located((By.ID, 'ty_rapport_eval')))
                except:
                     print("----------------Temporary block on IP------------")
                     break
                    
                try:
                    second_link = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//li[@id='Find-3']/a")))
                    second_link.click()
                except:
                    driver.quit()
                    # break
                delay()
                WebDriverWait(driver,10).until(EC.presence_of_element_located((By.ID, 'ty_rapport_eval'))).click()
                search = WebDriverWait(driver,10).until(EC.presence_of_element_located((By.ID, 'NoCadastre')))
                search.send_keys(nums[count])
                # check if search button clicks
                # time.sleep(3000)
                delay()
                try:
                    search_btn = driver.find_element(By.XPATH,"//button[@id='btnSearch']")
                    search_btn.click()
                    delay()

                    try:
                        if driver.find_element(By.CLASS_NAME, 'ch-container').find_element(By.XPATH, '/html/body/form/div[2]/div/div/div[1]').text[2:] == 'Aucun immeuble ne correspond à vos critères de recherche.':
                            df = pd.DataFrame([[nums[count], 'NO DATA FOUND']], columns=['Lot', 'Adresse'])
                            print(output, df)
                            # out = pd.read_excel(output)
                            out = get_sheetcol_values(output, 1)[1:]
                            # out = out.loc[~out.index.duplicated(keep='first')]
                            # out.drop_duplicates(keep = 'first', inplace = True)
                            # new_df = pd.concat([out,df], axis=0)
                            # new_df.reset_index(inplace=True, drop=True)
                            # new_df.to_excel(output, index=False)
                            data = df.values.tolist()
                            update_rowss(output, data)

                            # book = load_workbook(output)
                            # options = {}
                            # options['strings_to_formulas'] = False
                            # options['strings_to_urls'] = False
                            # with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={"options": options}) as writer:
                            #     writer.book = book
                            #     writer.sheets = {ws.title: ws for ws in book.worksheets}

                            #     for sheetname in writer.sheets:
                            #         # df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
                            #         df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
                            count+=1
                                    # continue
                    except Exception as e:
                        print(e)
                        pass

                    driver.find_element(By.ID, 'Row1').find_elements(By.TAG_NAME, 'td')[1].find_element(By.TAG_NAME, 'a').click()
                except NoSuchElementException:
                    driver.quit()
                    break
                try:
                    WebDriverWait(driver,5).until(EC.presence_of_element_located((By.ID, 'mail_usager')))
                    # domainn = select_domain()
                    latesttt = current_alias_func()
                    # user_alias = f'{current_alias_func()}@{domainn}'
                    user_alias = f'{latesttt[0]}@gmail.com'
                    WebDriverWait(driver,10).until(EC.presence_of_element_located((By.ID, 'mail_usager'))).send_keys(user_alias)
                    driver.find_element(By.CLASS_NAME, 'modal-footer').find_element(By.CLASS_NAME, 'pull-right').find_element(By.TAG_NAME, 'button').click()
                    time.sleep(2)
                    # if domainn == 'eirufw.xyz':
                    #     emailCode = email_code(username1,password)
                    # elif domainn == 'pqiqef.xyz':
                    #     emailCode = email_code(username2,password)
                    # elif domainn == 'wkjfher.xyz':
                    #     emailCode = email_code(username3,password)
                    emailCode = email_code(user_alias, latesttt[1])
                    WebDriverWait(driver,10).until(EC.presence_of_element_located((By.ID, 'code'))).send_keys(emailCode)
                    driver.find_element(By.CLASS_NAME, 'modal-footer').find_element(By.CLASS_NAME, 'pull-right').find_element(By.TAG_NAME, 'button').click()
                    delay()
                    # driver.switch_to.window(driver.window_handles[1])
                    # driver.find_element_by_xpath('/html/body/form/div[1]/div/div/div[3]/div/div[3]/div/div[1]/div[2]/div/div/button').click()
                except Exception as e:
                    pass
                # detail_btn = driver.find_element_by_xpath("/html/body/form/div[1]/div/div/div[3]/div/div[3]/div/div[1]/div[2]/div/div/button")
                # detail_btn.click()
                try:
                    driver.find_element_by_class_name('btn-acceo').click()
                    
                except NoSuchElementException:
                    driver.quit()
                    break
                delay()
                driver.switch_to.window(driver.window_handles[1])
                delay()
                iframe = driver.find_element_by_name("Report_PDF_Zone")
                driver.switch_to.frame(iframe)
                delay()
                WebDriverWait(driver,10).until(EC.presence_of_all_elements_located((By.TAG_NAME, 'tr')))
                all_tr = driver.find_elements_by_tag_name('tr')
                print('------LOOP STARTED------')
                # for ind, row in enumerate(all_tr[20:]):
                #  time.sleep(4)
                vall, key_list = get_col_vlaues(all_tr, driver)
                if vall[0] == '':
                    driver.quit()
                    break
                vall.insert(0, nums[count])
                combined.append(vall)
                df = pd.DataFrame(combined, columns=key_list)
                print(output, df)
                # outer = pd.read_excel(output)
                out = get_sheetcol_values(output, 1)[1:]   
                # out_val = outer.values.tolist()
                # out_val.append(df.values.tolist()[0])
                # new_df = pd.DataFrame(out_val, columns=key_list)
                # new_df
                # new_df.to_excel(output, index=False)
                data = df.values.tolist()
                update_rowss(output, data)

                
                # df.to_excel('statics\excel\output.xlsx', index=False)
                # book = load_workbook(output)
                # options = {}
                # options['strings_to_formulas'] = False
                # options['strings_to_urls'] = False
                # book = load_workbook(output)
                # options = {}
                # options['strings_to_formulas'] = False
                # options['strings_to_urls'] = False
                # with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={"options": options}) as writer:
                #     writer.book = book
                #     writer.sheets = {ws.title: ws for ws in book.worksheets}

                #     for sheetname in writer.sheets:
                #         # df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
                #         df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
                driver.switch_to.default_content()
                driver.find_element(By.XPATH, '/html/body/form/div/div[2]/div[1]/div/div/button').click()
                delay()
                old_name = f'{os.getcwd()}\\statics\\pdfs{thread}\\role_evaluation.pdf'
                new_name = f'{os.getcwd()}\\statics\\pdfs{thread}\\{vall[0]}_{vall[1]}.pdf'
                try:
                    os.rename(old_name, new_name)
                except FileNotFoundError:
                    delay()
                    try:
                        os.rename(old_name, new_name)
                    except:
                        pass
                except FileExistsError:
                    print("File already Exists")
                    print("Removing existing file")
                    os.remove(new_name)
                    # rename it
                    os.rename(old_name, new_name)
                    print('Done renaming a file')
                count+=1
                  

                time.sleep(1)
                driver.close()
                time.sleep(1)
                driver.quit()
            except Exception as e:
                print(e)
                driver.quit()
                continue

    # append_df_to_excel('statics\excel\Copy of PDF extract fields and format.xlsx', df, sheet_name='Sheet1', header=None, index=False)

if __name__ == '__main__':
    # outputs = ['statics\excel\output33.xlsx']
    outputs = ['output33', 'output44']
    outputs_done = []
    for output in outputs:
        # out = pd.read_excel(output)
        out = get_sheetcol_values(output, 1)[1:]
        # print(out)
        outputs_done.extend(out)
    inputs = pd.read_excel('statics\excel\SJSR_Granby.xlsx')
    inputs = inputs['Lot'].tolist()
    splits_diff = [str(item) for item in inputs if str(item) not in outputs_done]
    splits = np.array_split(splits_diff, 2)
    splits = [list(arr) for arr in splits]
    
    # with ThreadPool(2) as p:
    #     p.starmap(main,zip(splits, outputs, list(range(1, (len(outputs)+1)))))

    with concurrent.futures.ProcessPoolExecutor(2) as executor:
        executor.map(main, splits, outputs, list(range(1, (len(outputs)+1))))

    